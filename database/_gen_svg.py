#!/usr/bin/env python3
# Generates one SVG table per Excel sheet — provenance documentation for the
# "The Wretched of the Earth" knowledge graph. Styled in the site palette.
import openpyxl, html, re, os

XLSX = "/Users/mariagriga/Downloads/The Wretched of the Earth_MG -3.xlsx"
OUT  = os.path.dirname(os.path.abspath(__file__))

# ── site palette ──
BG      = "#f4f1ea"   # milk
INK     = "#2a2620"
ACCENT  = "#2e5c8b"   # blue
ACCENT2 = "#4a7fa5"
BAND    = "#e3edf6"   # light blue section band
ROW_ALT = "#ece8df"
GRID    = "#cdc7ba"
MUTED   = "#7a766a"

FS      = 12.5        # body font size
HFS     = 13          # header font size
CHARW   = 7.05        # approx mono char width @ FS
LH      = 18          # line height
PAD_X   = 12
PAD_Y   = 11
MAXW    = 1500        # max table width
TITLE_H = 64

def wrap(text, maxchars):
    text = str(text)
    words = text.split(' ')
    lines, cur = [], ''
    for w in words:
        if len(w) > maxchars:               # hard-break very long tokens (URIs)
            if cur: lines.append(cur); cur=''
            while len(w) > maxchars:
                lines.append(w[:maxchars]); w = w[maxchars:]
            cur = w
            continue
        test = (cur + ' ' + w).strip()
        if len(test) <= maxchars:
            cur = test
        else:
            if cur: lines.append(cur)
            cur = w
    if cur: lines.append(cur)
    return lines or ['']

def esc(s): return html.escape(str(s))

def build_sheet(ws):
    rows = []
    for r in ws.iter_rows(values_only=True):
        cells = ['' if c is None else str(c).strip() for c in r]
        while cells and cells[-1] == '': cells.pop()
        if any(c != '' for c in cells):
            rows.append(cells)
    if not rows: return None
    ncol = max(len(r) for r in rows)
    rows = [r + ['']*(ncol-len(r)) for r in rows]

    # classify rows: 'header' (first), 'section' (single non-empty cell), 'data'
    kinds = []
    for i, r in enumerate(rows):
        nonempty = [c for c in r if c != '']
        if i == 0:
            kinds.append('header')
        elif len(nonempty) == 1 and ncol > 1:
            kinds.append('section')
        else:
            kinds.append('data')

    # column natural widths (chars), capped
    CAP = 58
    colmax = [1]*ncol
    for i, r in enumerate(rows):
        if kinds[i] == 'section': continue
        for j in range(ncol):
            colmax[j] = max(colmax[j], min(len(r[j]), CAP))
    colw = [int(c*CHARW + PAD_X*2) for c in colmax]
    total = sum(colw)
    if total > MAXW:                       # scale down proportionally
        scale = MAXW/total
        colw = [max(70, int(w*scale)) for w in colw]
    colchars = [max(6, int((w-PAD_X*2)/CHARW)) for w in colw]
    tot_w = sum(colw)
    xpos = [0]
    for w in colw: xpos.append(xpos[-1]+w)

    # row heights
    heights = []
    for i, r in enumerate(rows):
        if kinds[i] == 'section':
            heights.append(LH + PAD_Y)
        else:
            maxlines = 1
            for j in range(ncol):
                maxlines = max(maxlines, len(wrap(r[j], colchars[j])))
            heights.append(maxlines*LH + PAD_Y)
    ypos = [TITLE_H]
    for h in heights: ypos.append(ypos[-1]+h)
    tot_h = ypos[-1] + 16

    # ── emit svg ──
    s = []
    s.append(f'<svg xmlns="http://www.w3.org/2000/svg" width="{tot_w}" height="{tot_h}" '
             f'viewBox="0 0 {tot_w} {tot_h}" font-family="\'IBM Plex Mono\', monospace">')
    s.append(f'<rect width="{tot_w}" height="{tot_h}" fill="{BG}"/>')
    # title
    s.append(f'<text x="{PAD_X}" y="30" font-size="20" font-weight="700" fill="{ACCENT}">'
             f'{esc(ws.title)}</text>')
    s.append(f'<text x="{PAD_X}" y="48" font-size="10.5" fill="{MUTED}" letter-spacing="1.5">'
             f'THE WRETCHED OF THE EARTH · DATA PROVENANCE</text>')
    s.append(f'<line x1="{PAD_X}" y1="{TITLE_H-6}" x2="{tot_w-PAD_X}" y2="{TITLE_H-6}" '
             f'stroke="{ACCENT2}" stroke-width="1.5"/>')

    for i, r in enumerate(rows):
        y0, y1 = ypos[i], ypos[i+1]
        h = y1-y0
        if kinds[i] == 'header':
            s.append(f'<rect x="0" y="{y0}" width="{tot_w}" height="{h}" fill="{ACCENT}"/>')
            for j in range(ncol):
                tx = xpos[j]+PAD_X
                s.append(f'<text x="{tx}" y="{y0+LH+1}" font-size="{HFS}" font-weight="600" '
                         f'fill="#ffffff">{esc(r[j])}</text>')
        elif kinds[i] == 'section':
            label = next((c for c in r if c != ''), '')
            s.append(f'<rect x="0" y="{y0}" width="{tot_w}" height="{h}" fill="{BAND}"/>')
            s.append(f'<rect x="0" y="{y0}" width="4" height="{h}" fill="{ACCENT}"/>')
            s.append(f'<text x="{PAD_X}" y="{y0+LH}" font-size="{FS}" font-weight="600" '
                     f'fill="{ACCENT}">{esc(label)}</text>')
        else:
            if (i % 2) == 0:
                s.append(f'<rect x="0" y="{y0}" width="{tot_w}" height="{h}" fill="{ROW_ALT}"/>')
            for j in range(ncol):
                tx = xpos[j]+PAD_X
                lines = wrap(r[j], colchars[j])
                fill = ACCENT2 if (j == 0) else INK
                weight = '500' if (j == 0) else '400'
                for k, ln in enumerate(lines):
                    s.append(f'<text x="{tx}" y="{y0+LH-3+k*LH}" font-size="{FS}" '
                             f'font-weight="{weight}" fill="{fill}">{esc(ln)}</text>')
    # grid: column separators
    for j in range(1, ncol):
        s.append(f'<line x1="{xpos[j]}" y1="{TITLE_H}" x2="{xpos[j]}" y2="{ypos[-1]}" '
                 f'stroke="{GRID}" stroke-width="0.6"/>')
    # row separators
    for i in range(len(rows)+1):
        s.append(f'<line x1="0" y1="{ypos[i]}" x2="{tot_w}" y2="{ypos[i]}" '
                 f'stroke="{GRID}" stroke-width="0.6"/>')
    s.append(f'<rect x="0.5" y="{TITLE_H}" width="{tot_w-1}" height="{ypos[-1]-TITLE_H}" '
             f'fill="none" stroke="{ACCENT2}" stroke-width="1"/>')
    s.append('</svg>')
    return '\n'.join(s)

def safe(name):
    return re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')

wb = openpyxl.load_workbook(XLSX, data_only=True)
order = {n:i for i,n in enumerate(wb.sheetnames)}
for ws in wb.worksheets:
    svg = build_sheet(ws)
    if not svg: continue
    fn = f"{order[ws.title]+1:02d}_{safe(ws.title)}.svg"
    with open(os.path.join(OUT, fn), 'w', encoding='utf-8') as f:
        f.write(svg)
    print("wrote", fn)
